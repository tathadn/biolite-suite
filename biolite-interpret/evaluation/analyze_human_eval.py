#!/usr/bin/env python3
"""
analyze_human_eval.py — Compute inter-annotator agreement and
LLM-as-Judge correlation from returned evaluation spreadsheets.

Usage:
    python analyze_human_eval.py \
        --eval_dir evaluation/human_eval/ \
        --judge_scores evaluation/results/judge_finetuned_3b_v2.json \
        --output evaluation/human_eval/analysis_results.json

Reports:
    1. Fleiss' kappa (3+ raters) or Cohen's kappa (2 raters) per criterion
    2. Pearson/Spearman correlation between human mean and LLM judge scores
    3. Per-criterion summary statistics
    4. Disagreement analysis (examples with high variance)
"""

import argparse
import glob
import json
import os
from collections import defaultdict

import numpy as np
from openpyxl import load_workbook
from scipy import stats


def load_evaluator_scores(xlsx_path):
    """Load scores from a returned evaluator spreadsheet."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Evaluation"]

    scores = []
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        ex_id, _, _, bio_acc, completeness, clarity, notes = row
        if ex_id is None:
            break
        scores.append({
            "id": str(ex_id),
            "biological_accuracy": int(bio_acc) if bio_acc else None,
            "completeness": int(completeness) if completeness else None,
            "clarity": int(clarity) if clarity else None,
            "notes": str(notes) if notes else "",
        })

    return scores


def cohens_kappa(rater1, rater2):
    """Compute Cohen's kappa for two raters."""
    assert len(rater1) == len(rater2)
    n = len(rater1)

    # Categories: 1-5
    categories = list(range(1, 6))
    k = len(categories)

    # Build confusion matrix
    matrix = np.zeros((k, k), dtype=int)
    for a, b in zip(rater1, rater2):
        if a is not None and b is not None:
            matrix[a - 1][b - 1] += 1

    n_valid = matrix.sum()
    if n_valid == 0:
        return 0.0

    # Observed agreement
    p_o = np.trace(matrix) / n_valid

    # Expected agreement
    row_sums = matrix.sum(axis=1) / n_valid
    col_sums = matrix.sum(axis=0) / n_valid
    p_e = np.sum(row_sums * col_sums)

    if p_e == 1.0:
        return 1.0

    return (p_o - p_e) / (1 - p_e)


def fleiss_kappa(ratings_matrix):
    """Compute Fleiss' kappa for 3+ raters.

    ratings_matrix: numpy array of shape (n_subjects, n_categories)
    where each cell is the number of raters who assigned that category.
    """
    n_subjects, n_categories = ratings_matrix.shape
    n_raters = ratings_matrix[0].sum()

    # Proportion of assignments to each category
    p_j = ratings_matrix.sum(axis=0) / (n_subjects * n_raters)

    # Per-subject agreement
    P_i = (np.sum(ratings_matrix ** 2, axis=1) - n_raters) / (n_raters * (n_raters - 1))

    P_bar = np.mean(P_i)
    P_e = np.sum(p_j ** 2)

    if P_e == 1.0:
        return 1.0

    return (P_bar - P_e) / (1 - P_e)


def build_fleiss_matrix(all_scores, criterion, n_categories=5):
    """Build the Fleiss kappa input matrix from evaluator scores."""
    n_examples = len(all_scores[0])
    n_raters = len(all_scores)

    matrix = np.zeros((n_examples, n_categories), dtype=int)

    for rater_scores in all_scores:
        for i, score_dict in enumerate(rater_scores):
            val = score_dict.get(criterion)
            if val is not None and 1 <= val <= 5:
                matrix[i][val - 1] += 1

    return matrix


def compute_correlations(human_means, judge_scores, criterion):
    """Compute Pearson and Spearman correlation between human and judge scores."""
    human_vals = []
    judge_vals = []

    for ex_id, h_mean in human_means.items():
        if ex_id in judge_scores:
            j_score = judge_scores[ex_id].get(criterion, {})
            if isinstance(j_score, dict):
                j_val = j_score.get("score")
            else:
                j_val = j_score
            if j_val is not None and h_mean is not None:
                human_vals.append(h_mean)
                judge_vals.append(j_val)

    if len(human_vals) < 3:
        return {"pearson": None, "spearman": None, "n": len(human_vals)}

    pearson_r, pearson_p = stats.pearsonr(human_vals, judge_vals)
    spearman_r, spearman_p = stats.spearmanr(human_vals, judge_vals)

    return {
        "pearson_r": round(pearson_r, 3),
        "pearson_p": round(pearson_p, 4),
        "spearman_r": round(spearman_r, 3),
        "spearman_p": round(spearman_p, 4),
        "n": len(human_vals),
    }


def kappa_interpretation(kappa):
    """Interpret kappa value."""
    if kappa < 0:
        return "poor"
    elif kappa < 0.20:
        return "slight"
    elif kappa < 0.40:
        return "fair"
    elif kappa < 0.60:
        return "moderate"
    elif kappa < 0.80:
        return "substantial"
    else:
        return "almost perfect"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--eval_dir", type=str, default="evaluation/human_eval/")
    parser.add_argument("--judge_scores", type=str, default=None,
                        help="Path to LLM judge scores for correlation analysis")
    parser.add_argument("--output", type=str, default="evaluation/human_eval/analysis_results.json")
    args = parser.parse_args()

    # Find evaluator files
    eval_files = sorted(glob.glob(os.path.join(args.eval_dir, "evaluator_*.xlsx")))
    print(f"Found {len(eval_files)} evaluator files")

    if len(eval_files) < 2:
        print("ERROR: Need at least 2 evaluator files for agreement analysis")
        return

    # Load all scores
    all_scores = []
    for f in eval_files:
        scores = load_evaluator_scores(f)
        all_scores.append(scores)
        print(f"  {os.path.basename(f)}: {len(scores)} examples, "
              f"{sum(1 for s in scores if s['biological_accuracy'] is not None)} rated")

    n_raters = len(all_scores)
    n_examples = len(all_scores[0])
    criteria = ["biological_accuracy", "completeness", "clarity"]

    results = {
        "n_raters": n_raters,
        "n_examples": n_examples,
        "criteria": {},
    }

    print(f"\n{'='*60}")
    print(f"Inter-Annotator Agreement ({n_raters} raters, {n_examples} examples)")
    print(f"{'='*60}")

    for criterion in criteria:
        # Compute mean scores per example
        human_means = {}
        human_stds = {}
        for i in range(n_examples):
            vals = [all_scores[r][i].get(criterion) for r in range(n_raters)]
            vals = [v for v in vals if v is not None]
            ex_id = all_scores[0][i]["id"]
            if vals:
                human_means[ex_id] = np.mean(vals)
                human_stds[ex_id] = np.std(vals)

        # Compute kappa
        if n_raters == 2:
            r1 = [s.get(criterion) for s in all_scores[0]]
            r2 = [s.get(criterion) for s in all_scores[1]]
            kappa = cohens_kappa(r1, r2)
            kappa_type = "cohen"
        else:
            matrix = build_fleiss_matrix(all_scores, criterion)
            kappa = fleiss_kappa(matrix)
            kappa_type = "fleiss"

        # Summary stats
        all_vals = []
        for r in range(n_raters):
            for s in all_scores[r]:
                v = s.get(criterion)
                if v is not None:
                    all_vals.append(v)

        # High-disagreement examples
        high_disagreement = [
            (ex_id, std) for ex_id, std in human_stds.items() if std > 1.0
        ]
        high_disagreement.sort(key=lambda x: -x[1])

        criterion_results = {
            "kappa": round(kappa, 3),
            "kappa_type": kappa_type,
            "kappa_interpretation": kappa_interpretation(kappa),
            "mean_score": round(np.mean(all_vals), 2),
            "std_score": round(np.std(all_vals), 2),
            "n_high_disagreement": len(high_disagreement),
        }

        # Correlation with LLM judge
        if args.judge_scores:
            with open(args.judge_scores) as f:
                judge_data = json.load(f)

            # Build judge score lookup by example ID
            judge_lookup = {}
            for result in judge_data.get("individual_results", []):
                idx = result.get("index", -1)
                if idx >= 0 and idx < n_examples:
                    ex_id = all_scores[0][idx]["id"]
                    judge_lookup[ex_id] = result.get("scores", {})

            corr = compute_correlations(human_means, judge_lookup, criterion)
            criterion_results["judge_correlation"] = corr

        results["criteria"][criterion] = criterion_results

        # Print
        interp = kappa_interpretation(kappa)
        print(f"\n{criterion.replace('_', ' ').title()}:")
        print(f"  {kappa_type.title()}'s kappa: {kappa:.3f} ({interp})")
        print(f"  Mean score: {np.mean(all_vals):.2f} ± {np.std(all_vals):.2f}")
        print(f"  High-disagreement examples (std > 1.0): {len(high_disagreement)}")
        if args.judge_scores and "judge_correlation" in criterion_results:
            corr = criterion_results["judge_correlation"]
            if corr.get("pearson_r") is not None:
                print(f"  vs LLM Judge — Pearson r={corr['pearson_r']:.3f} (p={corr['pearson_p']:.4f})")
                print(f"  vs LLM Judge — Spearman ρ={corr['spearman_r']:.3f} (p={corr['spearman_p']:.4f})")

    # Save results
    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "w") as f:
        json.dump(results, f, indent=2)

    print(f"\n{'='*60}")
    print(f"Results saved to {args.output}")

    # Print paper-ready summary
    print(f"\n{'='*60}")
    print("Paper-ready summary:")
    print(f"{'='*60}")
    for criterion in criteria:
        c = results["criteria"][criterion]
        k = c["kappa"]
        interp = c["kappa_interpretation"]
        print(f"  {criterion}: κ={k:.3f} ({interp}), M={c['mean_score']:.2f}±{c['std_score']:.2f}")


if __name__ == "__main__":
    main()
