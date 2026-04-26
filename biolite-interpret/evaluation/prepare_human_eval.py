#!/usr/bin/env python3
"""
prepare_human_eval.py — Select 30 stratified examples and generate
evaluation spreadsheets for biology PhD/Masters evaluators.

Reads model predictions, selects examples stratified by task_type,
randomizes order, and creates one Excel file per evaluator with
the rubric, calibration examples, and rating columns.

Usage:
    python prepare_human_eval.py \
        --predictions evaluation/results/finetuned_3b_v2_predictions.json \
        --test_data data/splits/test.json \
        --n_examples 30 \
        --n_evaluators 3 \
        --output_dir evaluation/human_eval/

Output:
    evaluation/human_eval/
        evaluator_1.xlsx
        evaluator_2.xlsx
        evaluator_3.xlsx
        selected_examples.json  (for later analysis)
"""

import argparse
import json
import os
import random

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# Rubric definitions
RUBRIC = {
    "biological_accuracy": {
        "description": "How biologically accurate is the interpretation?",
        "scale": {
            1: "Major factual errors — wrong gene functions, wrong pathways, hallucinated associations",
            2: "Multiple inaccuracies — some gene IDs correct but key biological claims are wrong",
            3: "Mostly correct — main themes right, but some imprecisions or minor errors",
            4: "Accurate — correct gene functions and pathway associations, minor issues only",
            5: "Fully accurate — all gene functions, directionalities, and pathway connections correct",
        },
    },
    "completeness": {
        "description": "Does the interpretation cover all important findings in the data?",
        "scale": {
            1: "Mentions <30% of key findings — skips major genes or ignores entire functional themes",
            2: "Covers some findings but misses important patterns",
            3: "Covers major findings but misses nuance or secondary themes",
            4: "Addresses most significant genes, enrichment themes, and caveats",
            5: "Comprehensive — all significant findings addressed with appropriate context",
        },
    },
    "clarity": {
        "description": "How clear and well-structured is the writing?",
        "scale": {
            1: "Jargon-heavy, incoherent, or disorganized — hard to follow",
            2: "Understandable but poorly structured or repetitive",
            3: "Adequate — readable but could be better organized",
            4: "Clear and well-structured — easy to follow for a biologist",
            5: "Excellent — well-organized, concise, accessible to biologist audience",
        },
    },
}

# Calibration examples (synthetic, not from test set)
CALIBRATION_EXAMPLES = [
    {
        "id": "CAL-1",
        "input": (
            "DESeq2 results from human breast cancer vs normal tissue:\n"
            "| Gene | log2FC | padj | baseMean |\n"
            "|------|--------|------|----------|\n"
            "| ESR1 | -3.21 | 1.2e-18 | 4521 |\n"
            "| ERBB2 | 4.87 | 3.4e-22 | 3102 |\n"
            "| MKI67 | 3.65 | 7.8e-15 | 2876 |\n"
            "| CDH1 | -2.43 | 4.5e-11 | 1987 |\n"
            "| TP53 | -1.89 | 2.1e-9 | 1654 |"
        ),
        "model_output": (
            "The differential expression profile reveals a classic aggressive breast "
            "cancer signature. ERBB2 (HER2) is strongly upregulated (log2FC=4.87), "
            "indicating HER2-positive subtype characteristics. The proliferation marker "
            "MKI67 (Ki-67) is similarly elevated, consistent with high mitotic activity. "
            "Downregulation of ESR1 (estrogen receptor alpha) suggests hormone receptor "
            "negativity, while loss of CDH1 (E-cadherin) indicates disrupted cell-cell "
            "adhesion and potential epithelial-mesenchymal transition. TP53 "
            "downregulation is consistent with p53 pathway inactivation. Together, "
            "these markers suggest a HER2+/ER- subtype with aggressive features "
            "including high proliferation and EMT activation."
        ),
        "suggested_scores": "bio_acc=5, completeness=5, clarity=5",
        "explanation": (
            "All gene identifications correct (ERBB2=HER2, MKI67=Ki-67, ESR1=ER-alpha, "
            "CDH1=E-cadherin). Directionalities match the table. Biological connections "
            "(HER2+/ER- subtype, EMT, p53 pathway) are accurate. All 5 genes addressed."
        ),
    },
    {
        "id": "CAL-2",
        "input": (
            "DESeq2 results from mouse liver, high-fat diet vs control:\n"
            "| Gene | log2FC | padj | baseMean |\n"
            "|------|--------|------|----------|\n"
            "| Fasn | 2.34 | 5.6e-12 | 3201 |\n"
            "| Pparg | 1.87 | 8.9e-10 | 1543 |\n"
            "| Cyp7a1 | -3.12 | 1.1e-15 | 2876 |\n"
            "| Acox1 | -1.95 | 3.4e-8 | 1234 |"
        ),
        "model_output": (
            "The results show activation of inflammatory pathways. Fasn upregulation "
            "indicates increased TNF-alpha signaling. Pparg is a transcription factor "
            "involved in cell cycle regulation and its upregulation suggests "
            "proliferative activity. Cyp7a1 downregulation reflects reduced kidney "
            "function. Acox1, a glycolytic enzyme, is suppressed indicating "
            "mitochondrial dysfunction."
        ),
        "suggested_scores": "bio_acc=1, completeness=2, clarity=3",
        "explanation": (
            "Major errors: Fasn is fatty acid synthase (lipogenesis), not TNF-alpha. "
            "PPARg is a lipid metabolism regulator, not cell cycle. Cyp7a1 is bile acid "
            "synthesis in LIVER, not kidney. Acox1 is peroxisomal fatty acid oxidation, "
            "not glycolysis. The real story is hepatic steatosis (lipogenesis up, "
            "beta-oxidation down). Writing is grammatically clear despite being wrong."
        ),
    },
]


def select_stratified_examples(test_data, predictions, n=30, seed=42):
    """Select n examples stratified by task_type."""
    random.seed(seed)

    # Match predictions to test data by index
    examples = []
    for i, (test_ex, pred) in enumerate(zip(test_data, predictions)):
        examples.append({
            "index": i,
            "input": pred.get("input", test_ex.get("input", "")),
            "instruction": test_ex.get("instruction", ""),
            "reference": pred.get("reference", test_ex.get("output", "")),
            "model_output": pred.get("prediction", ""),
            "task_type": test_ex.get("task_type",
                         test_ex.get("metadata", {}).get("task_type", "unknown")),
            "source": test_ex.get("source", "unknown"),
        })

    # Group by task type
    by_type = {}
    for ex in examples:
        tt = ex["task_type"]
        by_type.setdefault(tt, []).append(ex)

    # Proportional sampling
    selected = []
    total = len(examples)
    for tt, exs in by_type.items():
        proportion = len(exs) / total
        n_select = max(1, round(n * proportion))
        sampled = random.sample(exs, min(n_select, len(exs)))
        selected.extend(sampled)

    # Trim or pad to exactly n
    if len(selected) > n:
        selected = selected[:n]
    elif len(selected) < n:
        remaining = [e for e in examples if e not in selected]
        extra = random.sample(remaining, min(n - len(selected), len(remaining)))
        selected.extend(extra)

    # Randomize presentation order
    random.shuffle(selected)

    # Assign display IDs
    for i, ex in enumerate(selected):
        ex["display_id"] = f"EX-{i+1:02d}"

    return selected


def create_eval_spreadsheet(evaluator_id, examples, output_path):
    """Create a formatted evaluation spreadsheet for one evaluator."""
    wb = Workbook()

    # =========================================
    # Sheet 1: Instructions
    # =========================================
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr.sheet_properties.tabColor = "4472C4"

    header_font = Font(name="Arial", size=14, bold=True, color="002060")
    body_font = Font(name="Arial", size=11)
    bold_font = Font(name="Arial", size=11, bold=True)
    section_font = Font(name="Arial", size=12, bold=True, color="002060")

    ws_instr.column_dimensions["A"].width = 100

    instructions = [
        ("BioLite Model Evaluation", header_font),
        ("", body_font),
        ("Thank you for helping evaluate our bioinformatics interpretation model!", body_font),
        ("", body_font),
        ("WHAT YOU'LL DO:", section_font),
        ("You will review 30 model-generated interpretations of differential expression", body_font),
        ("and pathway enrichment results. For each interpretation, you'll rate three aspects:", body_font),
        ("", body_font),
        ("  1. Biological Accuracy — Are the gene functions and pathway associations correct?", body_font),
        ("  2. Completeness — Does it cover all the important findings in the data?", body_font),
        ("  3. Clarity — Is it well-written and easy to follow?", body_font),
        ("", body_font),
        ("HOW TO SCORE:", section_font),
        ("Each criterion uses a 1-5 scale. See the 'Rubric' sheet for detailed definitions.", body_font),
        ("Please also read the 'Calibration' sheet for two worked examples before starting.", body_font),
        ("", body_font),
        ("IMPORTANT NOTES:", section_font),
        ("• Judge the MODEL OUTPUT only — the reference answer is provided for context but", body_font),
        ("  you don't need to compare them directly.", body_font),
        ("• The model output may contain errors — that's what we're measuring.", body_font),
        ("• If you're unsure about a specific gene or pathway, score based on what you can", body_font),
        ("  verify. Use the 'Notes' column for uncertainty.", body_font),
        ("• No ML knowledge needed — we only need your biology expertise.", body_font),
        ("", body_font),
        ("ESTIMATED TIME: 45-60 minutes", section_font),
        ("", body_font),
        ("WORKFLOW:", section_font),
        ("1. Read 'Rubric' sheet (2 min)", body_font),
        ("2. Read 'Calibration' sheet — two worked examples with explanations (5 min)", body_font),
        ("3. Go to 'Evaluation' sheet and rate all 30 examples (40-50 min)", body_font),
        ("4. Save the file and send it back to Dave", body_font),
        ("", body_font),
        (f"Evaluator ID: {evaluator_id}", bold_font),
        ("Please do not discuss scores with other evaluators until all forms are returned.", body_font),
    ]

    for row_idx, (text, font) in enumerate(instructions, 1):
        cell = ws_instr.cell(row=row_idx, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True)

    # =========================================
    # Sheet 2: Rubric
    # =========================================
    ws_rubric = wb.create_sheet("Rubric")
    ws_rubric.sheet_properties.tabColor = "70AD47"

    ws_rubric.column_dimensions["A"].width = 25
    ws_rubric.column_dimensions["B"].width = 10
    ws_rubric.column_dimensions["C"].width = 80

    rubric_header_fill = PatternFill("solid", fgColor="002060")
    rubric_header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    criterion_fill = PatternFill("solid", fgColor="D6E4F0")
    criterion_font = Font(name="Arial", size=11, bold=True, color="002060")

    row = 1
    ws_rubric.cell(row=row, column=1, value="Evaluation Rubric").font = header_font
    row += 2

    for criterion, details in RUBRIC.items():
        # Criterion header
        for col in range(1, 4):
            cell = ws_rubric.cell(row=row, column=col)
            cell.fill = criterion_fill
        ws_rubric.cell(row=row, column=1, value=criterion.replace("_", " ").title()).font = criterion_font
        ws_rubric.cell(row=row, column=3, value=details["description"]).font = bold_font
        row += 1

        # Scale entries
        for score, desc in details["scale"].items():
            ws_rubric.cell(row=row, column=1, value="").font = body_font
            ws_rubric.cell(row=row, column=2, value=score).font = bold_font
            ws_rubric.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            ws_rubric.cell(row=row, column=3, value=desc).font = body_font
            ws_rubric.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
            row += 1

        row += 1  # gap between criteria

    # =========================================
    # Sheet 3: Calibration
    # =========================================
    ws_cal = wb.create_sheet("Calibration")
    ws_cal.sheet_properties.tabColor = "FFC000"

    ws_cal.column_dimensions["A"].width = 15
    ws_cal.column_dimensions["B"].width = 85

    row = 1
    ws_cal.cell(row=row, column=1, value="Calibration Examples").font = header_font
    row += 1
    ws_cal.cell(row=row, column=1, value="Review these two examples before starting. They show what good (5) and poor (1) outputs look like.").font = body_font
    row += 2

    for cal_ex in CALIBRATION_EXAMPLES:
        ws_cal.cell(row=row, column=1, value=f"Example {cal_ex['id']}").font = section_font
        row += 1

        fields = [
            ("Input:", cal_ex["input"]),
            ("Model Output:", cal_ex["model_output"]),
            ("Suggested Scores:", cal_ex["suggested_scores"]),
            ("Why:", cal_ex["explanation"]),
        ]
        for label, content in fields:
            ws_cal.cell(row=row, column=1, value=label).font = bold_font
            ws_cal.cell(row=row, column=2, value=content).font = body_font
            ws_cal.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1

        row += 2  # gap between calibration examples

    # =========================================
    # Sheet 4: Evaluation (main rating sheet)
    # =========================================
    ws_eval = wb.create_sheet("Evaluation")
    ws_eval.sheet_properties.tabColor = "FF0000"

    # Column widths
    col_widths = {
        "A": 8,   # ID
        "B": 60,  # Input
        "C": 60,  # Model Output
        "D": 14,  # Bio Accuracy
        "E": 14,  # Completeness
        "F": 14,  # Clarity
        "G": 40,  # Notes
    }
    for col_letter, width in col_widths.items():
        ws_eval.column_dimensions[col_letter].width = width

    # Headers
    headers = ["ID", "Input (DE/Enrichment Table)", "Model Output",
               "Biological\nAccuracy\n(1-5)", "Completeness\n(1-5)",
               "Clarity\n(1-5)", "Notes (optional)"]
    header_fill = PatternFill("solid", fgColor="002060")
    header_font_white = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws_eval.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws_eval.row_dimensions[1].height = 45

    # Score input styling
    score_fill = PatternFill("solid", fgColor="FFF2CC")
    input_font = Font(name="Arial", size=9)
    score_font = Font(name="Arial", size=12, bold=True)

    # Data rows
    for row_idx, ex in enumerate(examples, 2):
        # Alternate row coloring
        if row_idx % 2 == 0:
            row_fill = PatternFill("solid", fgColor="F2F2F2")
        else:
            row_fill = PatternFill("solid", fgColor="FFFFFF")

        # ID
        cell = ws_eval.cell(row=row_idx, column=1, value=ex["display_id"])
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.fill = row_fill
        cell.border = thin_border

        # Input
        input_text = ex.get("instruction", "")
        if ex.get("input"):
            input_text += "\n\n" + ex["input"]
        cell = ws_eval.cell(row=row_idx, column=2, value=input_text.strip())
        cell.font = input_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = row_fill
        cell.border = thin_border

        # Model output
        cell = ws_eval.cell(row=row_idx, column=3, value=ex["model_output"])
        cell.font = input_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = row_fill
        cell.border = thin_border

        # Score columns (highlighted for input)
        for col_idx in range(4, 7):
            cell = ws_eval.cell(row=row_idx, column=col_idx)
            cell.fill = score_fill
            cell.font = score_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Notes column
        cell = ws_eval.cell(row=row_idx, column=7)
        cell.fill = row_fill
        cell.font = input_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border

        # Row height
        ws_eval.row_dimensions[row_idx].height = 150

    # Freeze header row
    ws_eval.freeze_panes = "A2"

    # Set Evaluation as active sheet
    wb.active = wb.sheetnames.index("Evaluation")

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Prepare human evaluation spreadsheets for BioLite"
    )
    parser.add_argument("--predictions", type=str, required=True,
                        help="Path to model predictions JSON")
    parser.add_argument("--test_data", type=str, required=True,
                        help="Path to test split JSON")
    parser.add_argument("--n_examples", type=int, default=30)
    parser.add_argument("--n_evaluators", type=int, default=3)
    parser.add_argument("--output_dir", type=str, default="evaluation/human_eval/")
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    # Load data
    with open(args.predictions) as f:
        predictions = json.load(f)
    with open(args.test_data) as f:
        test_data = json.load(f)

    print(f"Loaded {len(predictions)} predictions, {len(test_data)} test examples")

    # Select examples
    selected = select_stratified_examples(
        test_data, predictions, n=args.n_examples, seed=args.seed
    )

    print(f"Selected {len(selected)} examples")
    task_dist = {}
    for ex in selected:
        tt = ex["task_type"]
        task_dist[tt] = task_dist.get(tt, 0) + 1
    print(f"Task type distribution: {task_dist}")

    # Save selected examples for later analysis
    selected_path = os.path.join(args.output_dir, "selected_examples.json")
    with open(selected_path, "w") as f:
        json.dump(selected, f, indent=2)
    print(f"Saved selected examples to {selected_path}")

    # Generate one spreadsheet per evaluator (same examples, same order)
    for i in range(1, args.n_evaluators + 1):
        evaluator_id = f"Evaluator_{i}"
        output_path = os.path.join(args.output_dir, f"evaluator_{i}.xlsx")
        create_eval_spreadsheet(evaluator_id, selected, output_path)
        print(f"Created {output_path}")

    print(f"\nDone! {args.n_evaluators} spreadsheets created in {args.output_dir}")
    print("Send one spreadsheet to each evaluator along with these instructions:")
    print("1. Open the file in Excel or Google Sheets")
    print("2. Read the 'Instructions' tab first")
    print("3. Read the 'Calibration' tab for worked examples")
    print("4. Rate all 30 examples in the 'Evaluation' tab")
    print("5. Save and return the file")


if __name__ == "__main__":
    main()
